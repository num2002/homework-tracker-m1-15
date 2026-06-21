from __future__ import annotations

import os
import tempfile
import hashlib
import hmac
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path

import gspread
import pandas as pd
import streamlit as st
from google.oauth2.service_account import Credentials
from gspread.exceptions import SpreadsheetNotFound, WorksheetNotFound
from streamlit.errors import StreamlitSecretNotFoundError
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


APP_DIR = Path(__file__).resolve().parent
EXCEL_FILE = APP_DIR / "Homework_Tracker_M1_15.xlsx"
DEFAULT_GOOGLE_SHEET_ID = "1JnxohQh-anm6HFJZudMOLHvnDl8osa-j2V7ynQhYQb0"
STATUSES = ["ยังไม่ส่ง", "ส่งแล้ว", "ส่งช้า", "ยกเว้น"]
STUDENT_COLUMNS = ["เลขที่", "รหัสนักเรียน", "ชื่อ-นามสกุล"]
HOMEWORK_COLUMNS = ["HW_ID", "วิชา", "รายละเอียดงาน", "วันที่สั่ง", "กำหนดส่ง", "หมายเหตุ"]
TRACKING_COLUMNS = ["HW_ID", "เลขที่", "รหัสนักเรียน", "ชื่อ-นามสกุล", "สถานะ", "วันที่ส่ง", "หมายเหตุ"]
SHEETS = {
    "Students": (STUDENT_COLUMNS, 3),
    "Homework": (HOMEWORK_COLUMNS, 6),
    "Tracking": (TRACKING_COLUMNS, 7),
}


def anonymous_students() -> pd.DataFrame:
    """Return the fixed anonymous roster (student numbers 1–40 only)."""
    return pd.DataFrame([[number, "", ""] for number in range(1, 41)], columns=STUDENT_COLUMNS)


def anonymize_frames(
    students: pd.DataFrame, homework: pd.DataFrame, tracking: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Remove identifying student data while retaining the legacy sheet columns."""
    students = anonymous_students()
    tracking = tracking.reindex(columns=TRACKING_COLUMNS).copy()
    if not tracking.empty:
        tracking["รหัสนักเรียน"] = ""
        tracking["ชื่อ-นามสกุล"] = ""
    return students, homework, tracking


def _style_sheet(ws, color: str, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    fill = PatternFill("solid", fgColor=color)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.auto_filter.ref = ws.dimensions


def _normalize_workbook(path: Path) -> None:
    """Add the required sheets/columns while preserving an existing workbook."""
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    definitions = {
        "Students": (STUDENT_COLUMNS, "2563EB", [10, 18, 34]),
        "Homework": (HOMEWORK_COLUMNS, "0F766E", [16, 20, 42, 16, 16, 30]),
        "Tracking": (TRACKING_COLUMNS, "7C3AED", [16, 10, 18, 34, 16, 16, 30]),
    }
    for name, (headers, color, widths) in definitions.items():
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        current = [ws.cell(1, i).value for i in range(1, len(headers) + 1)]
        if not any(current):
            for i, header in enumerate(headers, 1):
                ws.cell(1, i, header)
        elif current != headers:
            missing = [h for h in headers if h not in current]
            for header in missing:
                ws.cell(1, ws.max_column + 1, header)
        _style_sheet(ws, color, widths)

    students = wb["Students"]
    if students.max_row <= 1:
        for number in range(1, 41):
            students.append([number, "", ""])

    homework = wb["Homework"]
    for col in (4, 5):
        for row in range(2, max(homework.max_row, 2) + 1):
            homework.cell(row, col).number_format = "yyyy-mm-dd"
    tracking = wb["Tracking"]
    has_status_validation = any(str(item.sqref) == "E2:E10000" for item in tracking.data_validations.dataValidation)
    if not has_status_validation:
        validation = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"', allow_blank=False)
        tracking.add_data_validation(validation)
        validation.add("E2:E10000")
    if len(tracking.conditional_formatting) == 0:
        tracking.conditional_formatting.add("E2:E10000", FormulaRule(formula=['E2="ยังไม่ส่ง"'], fill=PatternFill("solid", fgColor="FEE2E2")))
        tracking.conditional_formatting.add("E2:E10000", FormulaRule(formula=['OR(E2="ส่งแล้ว",E2="ส่งช้า")'], fill=PatternFill("solid", fgColor="DCFCE7")))
    _atomic_save(wb, path)


def _atomic_save(wb, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix="homework_", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    try:
        wb.save(temp_name)
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def load_excel_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    _normalize_workbook(EXCEL_FILE)
    try:
        students = pd.read_excel(EXCEL_FILE, sheet_name="Students", dtype={"รหัสนักเรียน": str})
        homework = pd.read_excel(EXCEL_FILE, sheet_name="Homework", dtype={"HW_ID": str})
        tracking = pd.read_excel(EXCEL_FILE, sheet_name="Tracking", dtype={"HW_ID": str, "รหัสนักเรียน": str})
    except PermissionError as exc:
        raise RuntimeError("กรุณาปิดไฟล์ Excel ก่อนบันทึกหรือรีเฟรชข้อมูล") from exc
    students = students.reindex(columns=STUDENT_COLUMNS).fillna("")
    homework = homework.reindex(columns=HOMEWORK_COLUMNS)
    tracking = tracking.reindex(columns=TRACKING_COLUMNS)
    for frame, columns in ((homework, ["วันที่สั่ง", "กำหนดส่ง"]), (tracking, ["วันที่ส่ง"])):
        for column in columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce")
    if not tracking.empty:
        tracking["สถานะ"] = tracking["สถานะ"].fillna("ยังไม่ส่ง")
    return students, homework, tracking


def save_excel_frames(students: pd.DataFrame, homework: pd.DataFrame, tracking: pd.DataFrame) -> None:
    _normalize_workbook(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    for sheet_name, frame, columns in (
        ("Students", students, STUDENT_COLUMNS),
        ("Homework", homework, HOMEWORK_COLUMNS),
        ("Tracking", tracking, TRACKING_COLUMNS),
    ):
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        clean = frame.reindex(columns=columns).copy()
        clean = clean.where(pd.notna(clean), None)
        for row in clean.itertuples(index=False, name=None):
            ws.append([value.to_pydatetime() if isinstance(value, pd.Timestamp) else value for value in row])
    _style_sheet(wb["Students"], "2563EB", [10, 18, 34])
    _style_sheet(wb["Homework"], "0F766E", [16, 20, 42, 16, 16, 30])
    _style_sheet(wb["Tracking"], "7C3AED", [16, 10, 18, 34, 16, 16, 30])
    _atomic_save(wb, EXCEL_FILE)


def secret_value(key: str, default=None):
    try:
        return st.secrets[key]
    except (KeyError, FileNotFoundError, StreamlitSecretNotFoundError):
        return default


def google_sheets_configured() -> bool:
    return secret_value("gcp_service_account") is not None and bool(secret_value("GOOGLE_SHEET_ID", DEFAULT_GOOGLE_SHEET_ID))


@st.cache_resource(show_spinner=False)
def get_google_spreadsheet():
    info = dict(secret_value("gcp_service_account"))
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_info(info, scopes=scopes)
    client = gspread.authorize(credentials)
    try:
        return client.open_by_key(secret_value("GOOGLE_SHEET_ID", DEFAULT_GOOGLE_SHEET_ID))
    except SpreadsheetNotFound as exc:
        raise RuntimeError("เปิด Google Sheet ไม่ได้ กรุณาตรวจ GOOGLE_SHEET_ID และแชร์ชีตให้อีเมล Service Account") from exc


def _ensure_google_worksheet(spreadsheet, name: str, columns: list[str], min_rows: int):
    try:
        worksheet = spreadsheet.worksheet(name)
    except WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=name, rows=min_rows, cols=len(columns))
    values = worksheet.get_all_values()
    if not values:
        worksheet.update([columns], "A1")
    elif values[0][: len(columns)] != columns:
        worksheet.update([columns], "A1")
    return worksheet


def _google_values_to_frame(values: list[list[str]], columns: list[str]) -> pd.DataFrame:
    if len(values) <= 1:
        return pd.DataFrame(columns=columns)
    rows = [row[: len(columns)] + [""] * max(0, len(columns) - len(row)) for row in values[1:]]
    return pd.DataFrame(rows, columns=columns)


def load_google_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    spreadsheet = get_google_spreadsheet()
    frames = {}
    for name, (columns, column_count) in SHEETS.items():
        min_rows = 200 if name != "Tracking" else 10000
        worksheet = _ensure_google_worksheet(spreadsheet, name, columns, min_rows)
        frames[name] = _google_values_to_frame(worksheet.get_all_values(), columns)

    # On the first cloud launch, migrate the bundled Excel template/data automatically.
    if all(frame.empty for frame in frames.values()) and EXCEL_FILE.exists():
        excel_students, excel_homework, excel_tracking = load_excel_data()
        frames = {"Students": excel_students, "Homework": excel_homework, "Tracking": excel_tracking}
        for name, frame, columns in (
            ("Students", excel_students, STUDENT_COLUMNS),
            ("Homework", excel_homework, HOMEWORK_COLUMNS),
            ("Tracking", excel_tracking, TRACKING_COLUMNS),
        ):
            _write_google_worksheet(spreadsheet.worksheet(name), frame, columns)

    students = frames["Students"]
    if students.empty:
        students = anonymous_students()
        _write_google_worksheet(spreadsheet.worksheet("Students"), students, STUDENT_COLUMNS)
    students["เลขที่"] = pd.to_numeric(students["เลขที่"], errors="coerce")
    homework = frames["Homework"]
    tracking = frames["Tracking"]
    for frame, columns in ((homework, ["วันที่สั่ง", "กำหนดส่ง"]), (tracking, ["วันที่ส่ง"])):
        for column in columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce")
    if not tracking.empty:
        tracking["เลขที่"] = pd.to_numeric(tracking["เลขที่"], errors="coerce")
        tracking["สถานะ"] = tracking["สถานะ"].replace("", "ยังไม่ส่ง").fillna("ยังไม่ส่ง")
    return students.fillna(""), homework, tracking


def _value_for_google(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    if hasattr(value, "item"):
        value = value.item()
    return value


def _write_google_worksheet(worksheet, frame: pd.DataFrame, columns: list[str]) -> None:
    clean = frame.reindex(columns=columns)
    values = [columns] + [[_value_for_google(value) for value in row] for row in clean.itertuples(index=False, name=None)]
    worksheet.clear()
    worksheet.update(values, "A1")
    worksheet.freeze(rows=1)
    worksheet.format("1:1", {"backgroundColor": {"red": 0.12, "green": 0.36, "blue": 0.72}, "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True}, "horizontalAlignment": "CENTER"})


def save_google_frames(students: pd.DataFrame, homework: pd.DataFrame, tracking: pd.DataFrame) -> None:
    spreadsheet = get_google_spreadsheet()
    for name, frame, columns in (
        ("Students", students, STUDENT_COLUMNS),
        ("Homework", homework, HOMEWORK_COLUMNS),
        ("Tracking", tracking, TRACKING_COLUMNS),
    ):
        try:
            worksheet = spreadsheet.worksheet(name)
        except WorksheetNotFound:
            min_rows = max(len(frame) + 20, 200 if name != "Tracking" else 10000)
            worksheet = _ensure_google_worksheet(spreadsheet, name, columns, min_rows)
        _write_google_worksheet(worksheet, frame, columns)


@st.cache_data(ttl=15, show_spinner=False)
def load_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if google_sheets_configured():
        return anonymize_frames(*load_google_data())
    return anonymize_frames(*load_excel_data())


def save_frames(students: pd.DataFrame, homework: pd.DataFrame, tracking: pd.DataFrame) -> None:
    if google_sheets_configured():
        save_google_frames(students, homework, tracking)
    else:
        save_excel_frames(students, homework, tracking)
    load_data.clear()


def frames_to_excel(students: pd.DataFrame, homework: pd.DataFrame, tracking: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        students.to_excel(writer, sheet_name="Students", index=False)
        homework.to_excel(writer, sheet_name="Homework", index=False)
        tracking.to_excel(writer, sheet_name="Tracking", index=False)
    return buffer.getvalue()


def committee_accounts(master_password: str) -> list[dict[str, str]]:
    accounts = []
    for number in range(1, 6):
        username = f"committee{number}"
        digest = hmac.new(
            master_password.encode("utf-8"),
            f"m115-parent-network-{number}".encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        accounts.append(
            {
                "username": username,
                "password": f"M115-{digest[:10]}",
                "display_name": f"กรรมการผู้ปกครอง {number}",
            }
        )
    return accounts


def student_pin(master_password: str, student_number: int) -> str:
    """Create a stable PIN without storing student names or PINs in Google Sheets."""
    digest = hmac.new(
        master_password.encode("utf-8"),
        f"m115-student-{student_number}".encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"P{student_number:02d}-{digest[:6].upper()}"


def student_pin_accounts(master_password: str) -> list[dict[str, str | int]]:
    return [
        {"เลขที่": number, "PIN ผู้ปกครอง": student_pin(master_password, number)}
        for number in range(1, 41)
    ]


def append_audit(action: str, detail: str = "", actor: str | None = None) -> None:
    if not google_sheets_configured():
        return
    spreadsheet = get_google_spreadsheet()
    try:
        worksheet = spreadsheet.worksheet("Audit_Log")
    except WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title="Audit_Log", rows=1000, cols=4)
        worksheet.append_row(["วันที่เวลา", "ผู้ใช้", "รายการ", "รายละเอียด"])
    bangkok_time = datetime.utcnow() + timedelta(hours=7)
    worksheet.append_row(
        [
            bangkok_time.strftime("%Y-%m-%d %H:%M:%S"),
            actor or st.session_state.get("editor_name", "ไม่ทราบผู้ใช้"),
            action,
            detail,
        ]
    )


def submit_homework_for_student(
    students: pd.DataFrame,
    homework: pd.DataFrame,
    tracking: pd.DataFrame,
    hw_id: str,
    student_number: int,
) -> None:
    """Update one student's submission with a narrow Google Sheets write."""
    hw_rows = homework[homework["HW_ID"].astype(str) == str(hw_id)]
    if hw_rows.empty:
        raise ValueError("ไม่พบการบ้านที่เลือก")
    due_date = pd.to_datetime(hw_rows.iloc[0]["กำหนดส่ง"], errors="coerce")
    submitted_at = pd.Timestamp(date.today())
    status = "ส่งช้า" if pd.notna(due_date) and submitted_at > due_date else "ส่งแล้ว"

    if google_sheets_configured():
        worksheet = get_google_spreadsheet().worksheet("Tracking")
        values = worksheet.get_all_values()
        target_row = None
        for row_number, row in enumerate(values[1:], 2):
            row_hw_id = row[0].strip() if len(row) > 0 else ""
            try:
                row_student_number = int(float(row[1])) if len(row) > 1 and row[1] else None
            except ValueError:
                row_student_number = None
            if row_hw_id == str(hw_id) and row_student_number == student_number:
                target_row = row_number
                break
        if target_row is None:
            worksheet.append_row(
                [str(hw_id), student_number, "", "", status, submitted_at.strftime("%Y-%m-%d"), "ผู้ปกครองแจ้งส่ง"]
            )
        else:
            worksheet.update(
                [[status, submitted_at.strftime("%Y-%m-%d"), "ผู้ปกครองแจ้งส่ง"]],
                f"E{target_row}:G{target_row}",
            )
    else:
        mask = (
            (tracking["HW_ID"].astype(str) == str(hw_id))
            & (pd.to_numeric(tracking["เลขที่"], errors="coerce") == student_number)
        )
        if mask.any():
            tracking.loc[mask, ["สถานะ", "วันที่ส่ง", "หมายเหตุ"]] = [
                status,
                submitted_at,
                "ผู้ปกครองแจ้งส่ง",
            ]
        else:
            new_row = pd.DataFrame(
                [[str(hw_id), student_number, "", "", status, submitted_at, "ผู้ปกครองแจ้งส่ง"]],
                columns=TRACKING_COLUMNS,
            )
            tracking = pd.concat([tracking, new_row], ignore_index=True)
        save_frames(students, homework, tracking)

    load_data.clear()
    append_audit(
        "ผู้ปกครองแจ้งส่งการบ้าน",
        f"{hw_id} · เลขที่ {student_number} · {status}",
        actor=f"ผู้ปกครองเลขที่ {student_number}",
    )


def teacher_access() -> bool:
    configured_password = str(secret_value("APP_PASSWORD", "")).strip()
    if not configured_password:
        st.session_state.setdefault("editor_role", "teacher")
        st.session_state.setdefault("editor_name", "ครูผู้ดูแล")
        return True
    if st.session_state.get("authenticated"):
        st.sidebar.success(f"กำลังแก้ไขในชื่อ: {st.session_state.get('editor_name', 'ผู้ดูแล')}")
        if st.session_state.get("editor_role") == "teacher":
            credentials = pd.DataFrame(committee_accounts(configured_password))
            st.sidebar.download_button(
                "ดาวน์โหลดบัญชีกรรมการ 5 คน",
                credentials.to_csv(index=False).encode("utf-8-sig"),
                file_name="committee_accounts_m1_15.csv",
                mime="text/csv",
                use_container_width=True,
            )
            parent_pins = pd.DataFrame(student_pin_accounts(configured_password))
            st.sidebar.download_button(
                "ดาวน์โหลด PIN ผู้ปกครอง 40 เลขที่",
                parent_pins.to_csv(index=False).encode("utf-8-sig"),
                file_name="parent_pins_m1_15.csv",
                mime="text/csv",
                use_container_width=True,
            )
        if st.sidebar.button("ออกจากโหมดแก้ไข", use_container_width=True):
            st.session_state["authenticated"] = False
            st.session_state.pop("editor_role", None)
            st.session_state.pop("editor_name", None)
            st.rerun()
        return True
    with st.sidebar.expander("🔐 เข้าสู่ระบบครู/กรรมการ"):
        with st.form("teacher_login_form"):
            username = st.text_input("ชื่อผู้ใช้")
            password = st.text_input("รหัสผ่าน", type="password")
            login = st.form_submit_button("เข้าสู่ระบบ", type="primary", use_container_width=True)
        if login:
            normalized_username = username.strip().lower()
            account = next(
                (item for item in committee_accounts(configured_password) if item["username"] == normalized_username),
                None,
            )
            if normalized_username == "teacher" and hmac.compare_digest(password, configured_password):
                st.session_state["authenticated"] = True
                st.session_state["editor_role"] = "teacher"
                st.session_state["editor_name"] = "ครูผู้ดูแล"
                st.rerun()
            elif account and hmac.compare_digest(password, account["password"]):
                st.session_state["authenticated"] = True
                st.session_state["editor_role"] = "committee"
                st.session_state["editor_name"] = account["display_name"]
                st.rerun()
            else:
                st.error("ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง")
    return False


def next_hw_id(homework: pd.DataFrame) -> str:
    prefix = datetime.now().strftime("HW-%Y%m%d-")
    used = set(homework["HW_ID"].dropna().astype(str)) if not homework.empty else set()
    number = 1
    while f"{prefix}{number:03d}" in used:
        number += 1
    return f"{prefix}{number:03d}"


def thai_date(value) -> str:
    if pd.isna(value):
        return "-"
    dt = pd.Timestamp(value)
    months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    return f"{dt.day} {months[dt.month - 1]} {dt.year + 543}"


def homework_label(row: pd.Series) -> str:
    return f"{row['HW_ID']} · {row['วิชา']} · {row['รายละเอียดงาน']}"


st.set_page_config(
    page_title="Homework Tracker ม.1/15",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed",
)
st.markdown(
    """
<style>
  .block-container {padding-top: 1.2rem; padding-bottom: 3rem; max-width: 1200px;}
  [data-testid="stMetric"] {background:#f8fafc; border:1px solid #e2e8f0; padding:12px; border-radius:14px;}
  div[data-testid="stForm"] {border:1px solid #e2e8f0; border-radius:16px; padding:1rem;}
  @media (max-width: 640px) {.block-container {padding-left:.75rem; padding-right:.75rem;} h1 {font-size:1.65rem !important;}}
</style>
""",
    unsafe_allow_html=True,
)

is_editor = teacher_access()
st.title("📚 ติดตามการบ้าน ม.1/15")
st.caption("ใช้เฉพาะเลขที่ 1–40 · ไม่เก็บชื่อหรือรหัสนักเรียน")

try:
    students_df, homework_df, tracking_df = load_data()
except Exception as exc:
    st.error(f"เปิดข้อมูลไม่ได้: {exc}")
    st.stop()

storage_name = "Google Sheets ☁️" if google_sheets_configured() else f"Excel: {EXCEL_FILE.name}"
pin_secret = str(secret_value("APP_PASSWORD", "m115-local-parent-pin"))

if not is_editor:
    if not st.session_state.get("parent_authenticated"):
        st.subheader("เข้าสู่ระบบผู้ปกครอง")
        st.info("เลือกเลขที่ของบุตรหลานและใส่ PIN ที่ได้รับจากครูหรือกรรมการ")
        with st.form("parent_login_form"):
            selected_number = st.selectbox("เลขที่นักเรียน", range(1, 41))
            entered_pin = st.text_input("PIN ผู้ปกครอง", type="password")
            parent_login = st.form_submit_button("ดูการบ้านของเลขที่นี้", type="primary", use_container_width=True)
        if parent_login:
            expected_pin = student_pin(pin_secret, int(selected_number))
            if hmac.compare_digest(entered_pin.strip().upper(), expected_pin):
                st.session_state["parent_authenticated"] = True
                st.session_state["parent_number"] = int(selected_number)
                st.rerun()
            else:
                st.error("เลขที่หรือ PIN ไม่ถูกต้อง")
        st.stop()

    parent_number = int(st.session_state["parent_number"])
    st.sidebar.success(f"ผู้ปกครองเลขที่ {parent_number}")
    if st.sidebar.button("เปลี่ยนเลขที่ / ออกจากระบบ", use_container_width=True):
        st.session_state.pop("parent_authenticated", None)
        st.session_state.pop("parent_number", None)
        st.rerun()

    st.subheader(f"งานของนักเรียนเลขที่ {parent_number}")
    own_rows = tracking_df[
        pd.to_numeric(tracking_df["เลขที่"], errors="coerce") == parent_number
    ].copy()
    own_status = {
        str(row["HW_ID"]): row["สถานะ"] or "ยังไม่ส่ง"
        for _, row in own_rows.drop_duplicates("HW_ID", keep="last").iterrows()
    }
    pending_count = sum(
        own_status.get(str(row["HW_ID"]), "ยังไม่ส่ง") == "ยังไม่ส่ง"
        for _, row in homework_df.iterrows()
    )
    sent_count = sum(
        own_status.get(str(row["HW_ID"]), "ยังไม่ส่ง") in ["ส่งแล้ว", "ส่งช้า"]
        for _, row in homework_df.iterrows()
    )
    c1, c2, c3 = st.columns(3)
    c1.metric("งานทั้งหมด", len(homework_df))
    c2.metric("ค้างส่ง", pending_count)
    c3.metric("ส่งแล้ว", sent_count)

    if homework_df.empty:
        st.info("ยังไม่มีการบ้าน")
    else:
        ordered_homework = homework_df.sort_values("กำหนดส่ง", ascending=False, na_position="last")
        for _, hw in ordered_homework.iterrows():
            hw_id = str(hw["HW_ID"])
            status = own_status.get(hw_id, "ยังไม่ส่ง")
            icon = "⏳" if status == "ยังไม่ส่ง" else "✅" if status == "ส่งแล้ว" else "🕒" if status == "ส่งช้า" else "➖"
            with st.container(border=True):
                st.markdown(f"### {icon} {hw['วิชา']} — {status}")
                st.write(hw["รายละเอียดงาน"])
                st.caption(f"กำหนดส่ง: {thai_date(hw['กำหนดส่ง'])}")
                if str(hw.get("หมายเหตุ", "")).strip():
                    st.caption(f"หมายเหตุ: {hw['หมายเหตุ']}")
                if status == "ยังไม่ส่ง":
                    if st.button("แจ้งว่าส่งการบ้านแล้ว", key=f"parent-submit-{parent_number}-{hw_id}", type="primary", use_container_width=True):
                        try:
                            submit_homework_for_student(
                                students_df, homework_df, tracking_df, hw_id, parent_number
                            )
                            st.success("บันทึกการส่งแล้ว")
                            st.rerun()
                        except Exception as exc:
                            st.error(f"บันทึกไม่ได้: {exc}")
    st.stop()

pages = [
    "ภาพรวมทั้งห้อง",
    "เพิ่มการบ้าน",
    "แก้ไขการบ้าน",
    "อัปเดตสถานะส่งงาน",
    "สรุปข้อความส่ง LINE",
]
page = st.sidebar.radio("เมนูครู/กรรมการ", pages)
st.sidebar.caption(f"ฐานข้อมูล: {storage_name}")
if google_sheets_configured():
    st.sidebar.download_button(
        "ดาวน์โหลดข้อมูลสำรอง Excel",
        data=frames_to_excel(students_df, homework_df, tracking_df),
        file_name=f"Homework_Tracker_M1_15_backup_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

if page == "ภาพรวมทั้งห้อง":
    st.subheader("ภาพรวมทั้งห้องสำหรับครู/กรรมการ")
    today = pd.Timestamp(date.today())
    due = pd.to_datetime(homework_df["กำหนดส่ง"], errors="coerce") if not homework_df.empty else pd.Series(dtype="datetime64[ns]")
    near_due = int(((due >= today) & (due <= today + pd.Timedelta(days=7))).sum())
    c1, c2, c3 = st.columns(3)
    c1.metric("งานทั้งหมด", len(homework_df))
    c2.metric("ใกล้ครบกำหนด (7 วัน)", near_due)
    c3.metric("รายการค้างส่ง", int((tracking_df["สถานะ"] == "ยังไม่ส่ง").sum()) if not tracking_df.empty else 0)
    if homework_df.empty:
        st.info("ยังไม่มีการบ้าน")
    else:
        summary_rows = []
        for _, hw in homework_df.sort_values("กำหนดส่ง", na_position="last").iterrows():
            rows = tracking_df[tracking_df["HW_ID"].astype(str) == str(hw["HW_ID"])]
            sent = int(rows["สถานะ"].isin(["ส่งแล้ว", "ส่งช้า"]).sum())
            pending = rows[rows["สถานะ"] == "ยังไม่ส่ง"]
            pending_numbers = ", ".join(str(int(number)) for number in pd.to_numeric(pending["เลขที่"], errors="coerce").dropna())
            summary_rows.append({
                "HW_ID": hw["HW_ID"], "วิชา": hw["วิชา"], "งาน": hw["รายละเอียดงาน"],
                "กำหนดส่ง": hw["กำหนดส่ง"], "ส่งแล้ว": sent, "ค้างส่ง": len(pending),
                "เลขที่ค้างส่ง": pending_numbers or "-",
            })
        st.dataframe(
            pd.DataFrame(summary_rows), use_container_width=True, hide_index=True,
            column_config={"กำหนดส่ง": st.column_config.DateColumn(format="DD/MM/YYYY"), "เลขที่ค้างส่ง": st.column_config.TextColumn(width="large")},
        )

elif page == "เพิ่มการบ้าน":
    st.subheader("เพิ่มการบ้านใหม่")
    with st.form("add_homework", clear_on_submit=True):
        hw_id = st.text_input("HW_ID", value=next_hw_id(homework_df))
        subject = st.text_input("วิชา *")
        detail = st.text_area("รายละเอียดงาน *")
        c1, c2 = st.columns(2)
        assigned = c1.date_input("วันที่สั่ง", value=date.today())
        due_date = c2.date_input("กำหนดส่ง", value=date.today() + timedelta(days=7))
        note = st.text_area("หมายเหตุ")
        submitted = st.form_submit_button("เพิ่มการบ้านและสร้างรายการ 40 เลขที่", type="primary", use_container_width=True)
    if submitted:
        errors = []
        if not hw_id.strip() or not subject.strip() or not detail.strip():
            errors.append("กรุณากรอก HW_ID วิชา และรายละเอียดงาน")
        if hw_id.strip() in set(homework_df["HW_ID"].astype(str)):
            errors.append("HW_ID นี้มีอยู่แล้ว")
        if due_date < assigned:
            errors.append("กำหนดส่งต้องไม่ก่อนวันที่สั่ง")
        if errors:
            st.error(" · ".join(errors))
        else:
            new_hw = pd.DataFrame([[hw_id.strip(), subject.strip(), detail.strip(), pd.Timestamp(assigned), pd.Timestamp(due_date), note.strip()]], columns=HOMEWORK_COLUMNS)
            new_tracking = pd.DataFrame({
                "HW_ID": hw_id.strip(), "เลขที่": range(1, 41), "รหัสนักเรียน": "", "ชื่อ-นามสกุล": "",
                "สถานะ": "ยังไม่ส่ง", "วันที่ส่ง": pd.NaT, "หมายเหตุ": "",
            })
            try:
                save_frames(
                    students_df,
                    pd.concat([homework_df, new_hw], ignore_index=True),
                    pd.concat([tracking_df, new_tracking], ignore_index=True),
                )
                append_audit("เพิ่มการบ้าน", f"{hw_id.strip()} · {subject.strip()}")
                st.success(f"เพิ่ม {hw_id} และสร้างรายการติดตาม 40 เลขที่แล้ว")
                st.rerun()
            except Exception as exc:
                st.error(f"บันทึกไม่ได้: {exc}")

elif page == "แก้ไขการบ้าน":
    st.subheader("แก้ไขข้อความและกำหนดส่ง")
    if homework_df.empty:
        st.info("ยังไม่มีการบ้าน")
    else:
        options = {homework_label(row): str(row["HW_ID"]) for _, row in homework_df.iloc[::-1].iterrows()}
        selected = st.selectbox("เลือกการบ้านที่ต้องการแก้ไข", list(options), key="edit_hw_select")
        edit_hw_id = options[selected]
        hw_index = homework_df.index[homework_df["HW_ID"].astype(str) == edit_hw_id][0]
        hw = homework_df.loc[hw_index]
        assigned_value = pd.Timestamp(hw["วันที่สั่ง"]).date() if pd.notna(hw["วันที่สั่ง"]) else date.today()
        due_value = pd.Timestamp(hw["กำหนดส่ง"]).date() if pd.notna(hw["กำหนดส่ง"]) else date.today()
        with st.form(f"edit_homework_form_{edit_hw_id}"):
            st.text_input("HW_ID", value=edit_hw_id, disabled=True)
            subject = st.text_input("วิชา *", value=str(hw["วิชา"] or ""))
            detail = st.text_area("รายละเอียดงาน *", value=str(hw["รายละเอียดงาน"] or ""), height=140)
            c1, c2 = st.columns(2)
            assigned = c1.date_input("วันที่สั่ง", value=assigned_value)
            due_date = c2.date_input("กำหนดส่ง", value=due_value)
            note = st.text_area("หมายเหตุ", value=str(hw["หมายเหตุ"] or ""))
            save_homework = st.form_submit_button("บันทึกการแก้ไข", type="primary", use_container_width=True)
        if save_homework:
            if not subject.strip() or not detail.strip():
                st.error("กรุณากรอกวิชาและรายละเอียดงาน")
            elif due_date < assigned:
                st.error("กำหนดส่งต้องไม่ก่อนวันที่สั่ง")
            else:
                homework_df.loc[hw_index, HOMEWORK_COLUMNS[1:]] = [
                    subject.strip(), detail.strip(), pd.Timestamp(assigned), pd.Timestamp(due_date), note.strip()
                ]
                try:
                    save_frames(students_df, homework_df, tracking_df)
                    append_audit("แก้ไขการบ้าน", f"{edit_hw_id} · {subject.strip()}")
                    st.success("บันทึกข้อความการบ้านแล้ว")
                except Exception as exc:
                    st.error(f"บันทึกไม่ได้: {exc}")

elif page == "อัปเดตสถานะส่งงาน":
    st.subheader("อัปเดตสถานะเลขที่ 1–40")
    if homework_df.empty:
        st.info("ยังไม่มีการบ้าน")
    else:
        options = {homework_label(row): str(row["HW_ID"]) for _, row in homework_df.iloc[::-1].iterrows()}
        selected = st.selectbox("เลือกการบ้าน", list(options), key="status_hw_select")
        status_hw_id = options[selected]
        mask = tracking_df["HW_ID"].astype(str) == status_hw_id
        existing = tracking_df.loc[mask].copy()
        existing["เลขที่"] = pd.to_numeric(existing["เลขที่"], errors="coerce")
        existing = existing.dropna(subset=["เลขที่"]).drop_duplicates("เลขที่", keep="last").set_index("เลขที่")
        visible = pd.DataFrame({"เลขที่": range(1, 41)})
        visible["สถานะ"] = visible["เลขที่"].map(existing["สถานะ"] if "สถานะ" in existing else {}).fillna("ยังไม่ส่ง")
        visible["วันที่ส่ง"] = pd.to_datetime(visible["เลขที่"].map(existing["วันที่ส่ง"] if "วันที่ส่ง" in existing else {}), errors="coerce")
        visible["หมายเหตุ"] = visible["เลขที่"].map(existing["หมายเหตุ"] if "หมายเหตุ" in existing else {}).fillna("")
        edited = st.data_editor(
            visible,
            use_container_width=True,
            hide_index=True,
            disabled=["เลขที่"],
            column_config={
                "เลขที่": st.column_config.NumberColumn(format="%d", width="small"),
                "สถานะ": st.column_config.SelectboxColumn(options=STATUSES, required=True),
                "วันที่ส่ง": st.column_config.DateColumn(format="DD/MM/YYYY"),
                "หมายเหตุ": st.column_config.TextColumn(width="medium"),
            },
            key=f"status-editor-{status_hw_id}",
        )
        st.caption("แตะช่องสถานะของเลขที่ที่ต้องการ แล้วกดบันทึกด้านล่าง")
        if st.button("บันทึกสถานะทั้ง 40 เลขที่", type="primary", use_container_width=True):
            edited = edited.copy()
            edited["สถานะ"] = edited["สถานะ"].where(edited["สถานะ"].isin(STATUSES), "ยังไม่ส่ง")
            edited["วันที่ส่ง"] = pd.to_datetime(edited["วันที่ส่ง"], errors="coerce")
            sent_without_date = edited["สถานะ"].isin(["ส่งแล้ว", "ส่งช้า"]) & edited["วันที่ส่ง"].isna()
            edited.loc[sent_without_date, "วันที่ส่ง"] = pd.Timestamp(date.today())
            due_value = pd.to_datetime(
                homework_df.loc[homework_df["HW_ID"].astype(str) == status_hw_id, "กำหนดส่ง"].iloc[0],
                errors="coerce",
            )
            if pd.notna(due_value):
                auto_late = (edited["สถานะ"] == "ส่งแล้ว") & edited["วันที่ส่ง"].notna() & (edited["วันที่ส่ง"] > due_value)
                edited.loc[auto_late, "สถานะ"] = "ส่งช้า"
            replacement = pd.DataFrame({
                "HW_ID": status_hw_id,
                "เลขที่": edited["เลขที่"].astype(int),
                "รหัสนักเรียน": "",
                "ชื่อ-นามสกุล": "",
                "สถานะ": edited["สถานะ"],
                "วันที่ส่ง": edited["วันที่ส่ง"],
                "หมายเหตุ": edited["หมายเหตุ"].fillna(""),
            })
            try:
                tracking_df = pd.concat([tracking_df.loc[~mask], replacement], ignore_index=True)
                save_frames(students_df, homework_df, tracking_df)
                append_audit("อัปเดตสถานะ", status_hw_id)
                st.success("บันทึกสถานะทั้ง 40 เลขที่แล้ว")
            except Exception as exc:
                st.error(f"บันทึกไม่ได้: {exc}")

elif page == "สรุปข้อความส่ง LINE":
    st.subheader("สร้างข้อความสรุปสำหรับ LINE")
    if homework_df.empty:
        st.info("ยังไม่มีการบ้าน")
    else:
        options = {homework_label(row): str(row["HW_ID"]) for _, row in homework_df.iloc[::-1].iterrows()}
        selected = st.selectbox("เลือกการบ้าน", list(options), key="line_hw")
        line_hw_id = options[selected]
        hw = homework_df[homework_df["HW_ID"].astype(str) == line_hw_id].iloc[0]
        rows = tracking_df[tracking_df["HW_ID"].astype(str) == line_hw_id].sort_values("เลขที่")
        sent = int(rows["สถานะ"].isin(["ส่งแล้ว", "ส่งช้า"]).sum())
        pending = rows[rows["สถานะ"] == "ยังไม่ส่ง"]
        pending_lines = [f"{i}. เลขที่ {int(row['เลขที่'])}" for i, (_, row) in enumerate(pending.iterrows(), 1)]
        message = "\n".join([
            "สรุปการบ้าน ม.1/15", f"วิชา: {hw['วิชา']}", f"งาน: {hw['รายละเอียดงาน']}",
            f"กำหนดส่ง: {thai_date(hw['กำหนดส่ง'])}", "", f"ส่งแล้ว: {sent} คน",
            f"ค้างส่ง: {len(pending)} คน", "รายชื่อค้างส่ง:", *(pending_lines or ["ไม่มี 🎉"]),
        ])
        st.text_area("ข้อความพร้อมคัดลอก", value=message, height=360)
        st.download_button("ดาวน์โหลดเป็นไฟล์ข้อความ", message.encode("utf-8-sig"), file_name=f"LINE_{line_hw_id}.txt", mime="text/plain", use_container_width=True)
